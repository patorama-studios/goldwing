"""
Generate two import CSVs from the corrected AGA members Excel.
Run:  python3 scripts/data/generate_import_csvs.py

Output:
  scripts/data/import_main_life.csv   — FULL + LIFE members (import first)
  scripts/data/import_associates.csv  — ASSOCIATE members  (import second)
"""

import csv
import datetime
import openpyxl

EXCEL_PATH = 'scripts/data/AGA Members by Chapter — CORRECTED.xlsx'
OUT_MAIN   = 'scripts/data/import_main_life.csv'
OUT_ASSOC  = 'scripts/data/import_associates.csv'

SKIP_SHEETS = ('SUMMARY', 'Corrections Log')

FIELDNAMES = [
    'member_id', 'first_name', 'last_name', 'email', 'phone',
    'member_type', 'chapter', 'status',
    'address_line1', 'address_line2', 'suburb', 'state', 'postal_code', 'country',
    'privacy_level', 'exclude_electronic', 'is_historic',
    'full_member_number',
]

def clean(v):
    if v is None:
        return ''
    if isinstance(v, bool):
        return '1' if v else '0'
    if isinstance(v, datetime.datetime):
        return v.strftime('%Y-%m-%d')
    return str(v).strip()

def bool_flag(v):
    if isinstance(v, bool):
        return '1' if v else '0'
    s = str(v).strip().upper() if v else ''
    return '1' if s in ('1', 'TRUE', 'YES', 'Y') else '0'

def col_index(headers, *candidates):
    """Find the first matching column index from a list of candidate name fragments."""
    for cand in candidates:
        for i, h in enumerate(headers):
            if cand.lower() in h.lower():
                return i
    return None

wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

main_rows  = []
assoc_rows = []

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue

    ws = wb[sheet_name]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if len(all_rows) < 3:
        continue

    # Row 0 = chapter title banner, Row 1 = column headers, Row 2+ = data
    raw_headers = [str(h).strip() if h is not None else '' for h in all_rows[1]]

    # Locate column indices by partial name match
    c_member_no   = col_index(raw_headers, 'Member #', 'Member#')
    c_last_m      = col_index(raw_headers, 'Surname(M)')
    c_first_m     = col_index(raw_headers, 'First Name(M)')
    c_hist_m      = col_index(raw_headers, 'Historic(M)')
    c_last_a      = col_index(raw_headers, 'Surname(A)')
    c_first_a     = col_index(raw_headers, 'First Name(A)')
    c_hist_a      = col_index(raw_headers, 'Historic(A)')
    c_addr1       = col_index(raw_headers, 'Postal Address 1')
    c_addr2       = col_index(raw_headers, 'Postal Address 2')
    c_suburb      = col_index(raw_headers, 'Suburb')
    c_state       = col_index(raw_headers, 'State')
    c_postcode    = col_index(raw_headers, 'Postcode')
    c_phone_m     = col_index(raw_headers, 'Phone(M)(M)')
    c_phone_a     = col_index(raw_headers, 'Phone(M)(A)')
    c_email_m     = col_index(raw_headers, 'EMail(M)')
    c_email_a     = col_index(raw_headers, 'EMail(A)')
    c_chapter     = col_index(raw_headers, 'Chapter')
    c_dir_info    = col_index(raw_headers, 'Directory Info')
    c_life        = col_index(raw_headers, 'Life Member')
    c_edir        = col_index(raw_headers, 'eDirectory')

    def get(row, idx):
        if idx is None or idx >= len(row):
            return ''
        return row[idx]

    for row in all_rows[2:]:  # skip banner + header rows
        if not any(row):
            continue

        member_no_raw = clean(get(row, c_member_no))
        if not member_no_raw:
            continue

        try:
            base = int(float(member_no_raw))
        except (ValueError, TypeError):
            continue

        last_m  = clean(get(row, c_last_m))
        first_m = clean(get(row, c_first_m))
        if not last_m and not first_m:
            continue

        is_life     = bool_flag(get(row, c_life))
        member_type = 'LIFE' if is_life == '1' else 'FULL'
        hist_m      = bool_flag(get(row, c_hist_m))
        email_m     = clean(get(row, c_email_m))
        phone_m     = clean(get(row, c_phone_m))
        addr1       = clean(get(row, c_addr1))
        addr2       = clean(get(row, c_addr2))
        suburb      = clean(get(row, c_suburb))
        state       = clean(get(row, c_state))
        postcode    = clean(get(row, c_postcode))
        chapter     = clean(get(row, c_chapter)) or sheet_name
        dir_info    = clean(get(row, c_dir_info))
        # Source `eDirectory` column means "INCLUDE in electronic directory"
        # (True = show me). Our DB column `exclude_electronic` is the opposite
        # (1 = hide me). Invert so the polarity is right.
        edir_in     = bool_flag(get(row, c_edir))  # 1 = wants to be IN
        edir        = '0' if edir_in == '1' else '1'  # → exclude_electronic value

        # privacy_level: take the first letter from Directory Info (e.g. "A", "B — Name + Address")
        privacy = dir_info[:1].upper() if dir_info else 'A'
        if privacy not in ('A', 'B', 'C', 'D', 'E', 'F'):
            privacy = 'A'

        main_rows.append({
            'member_id':         str(base),
            'first_name':        first_m,
            'last_name':         last_m,
            'email':             email_m,
            'phone':             phone_m,
            'member_type':       member_type,
            'chapter':           chapter,
            'status':            'active',
            'address_line1':     addr1,
            'address_line2':     addr2,
            'suburb':            suburb,
            'state':             state,
            'postal_code':       postcode,
            'country':           'Australia',
            'privacy_level':     privacy,
            'exclude_electronic': edir,
            'is_historic':       hist_m,
            'full_member_number': '',
        })

        # Associate
        last_a  = clean(get(row, c_last_a))
        first_a = clean(get(row, c_first_a))
        if not last_a and not first_a:
            continue

        hist_a  = bool_flag(get(row, c_hist_a))
        email_a = clean(get(row, c_email_a))
        phone_a = clean(get(row, c_phone_a))

        # If associate shares the main member's email, leave blank — household login
        # means they don't need their own email to access the account.
        if email_a.lower() == email_m.lower():
            email_a = ''

        assoc_rows.append({
            'member_id':         f'{base}.1',
            'first_name':        first_a,
            'last_name':         last_a,
            'email':             email_a,
            'phone':             phone_a,
            'member_type':       'ASSOCIATE',
            'chapter':           chapter,
            'status':            'active',
            'address_line1':     addr1,
            'address_line2':     addr2,
            'suburb':            suburb,
            'state':             state,
            'postal_code':       postcode,
            'country':           'Australia',
            'privacy_level':     privacy,
            'exclude_electronic': edir,
            'is_historic':       hist_a,
            'full_member_number': str(base),
        })

wb.close()

def write_csv(path, rows):
    with open(path, 'w', newline='', encoding='utf-8') as f:
        w = csv.DictWriter(f, fieldnames=FIELDNAMES)
        w.writeheader()
        w.writerows(rows)
    print(f'Wrote {len(rows)} rows → {path}')

write_csv(OUT_MAIN,  main_rows)
write_csv(OUT_ASSOC, assoc_rows)

# Quick summary
no_email_assoc = [r for r in assoc_rows if not r['email']]
print(f'\nAssociates with no email (will import without user login): {len(no_email_assoc)}')
for r in no_email_assoc:
    print(f'  {r["member_id"]}  {r["first_name"]} {r["last_name"]}')
